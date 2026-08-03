"""Source-backed research for one durable application workspace."""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from difflib import SequenceMatcher
from io import BytesIO
from typing import Callable, Literal

import requests
from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from ats_terms import build_job_ats_terms
from job_precompute import display_salary
from job_visibility import apply_public_job_visibility
from models import ScrapedJob, TrackedJob


MOM_WAGE_SOURCE_URL = "https://stats.mom.gov.sg/iMAS_Tables1/Wages/Wages_2025/mrsd_2025Wages_table4.xlsx"
MOM_WAGE_LANDING_URL = "https://stats.mom.gov.sg/Pages/Occupational-Wages-Tables2025.aspx"
MOM_WAGE_DATA_PERIOD = "June 2025"
MOM_WAGE_RELEASE_DATE = "2026-06-30"
MOM_WAGE_SHEET = "T4"
MOM_WAGE_FIRST_DATA_ROW = 10
MOM_WAGE_TIMEOUT_SECONDS = 15
MOM_WAGE_MATCH_MIN_SCORE = 0.34
HAYS_SALARY_GUIDE_URL = "https://www.hays.com.sg/salary-guide"
HAYS_SALARY_GUIDE_PUBLICATION_DATE = "2026"
MICHAEL_PAGE_SALARY_BENCHMARK_URL = "https://www.michaelpage.com.sg/salary-benchmark-tool"
COMPARABLE_JOB_LIMIT = 8
COMPANY_JOB_LIMIT = 5
ATS_TERM_LIMIT = 12
INTERVIEW_TERM_LIMIT = 4
SOURCE_STALE_AFTER_DAYS = 30
MIN_EVIDENCE_LINE_CHARS = 24
MAX_EVIDENCE_LINE_CHARS = 700
TITLE_QUERY_TERM_LIMIT = 4
OCCUPATION_JACCARD_WEIGHT = 0.65
OCCUPATION_SEQUENCE_WEIGHT = 0.35
OCCUPATION_SUBSET_BONUS = 0.15
OCCUPATION_MATCH_SCORE_DECIMALS = 3
RECURRING_TERM_MIN_POSTINGS = 2
PRIMARY_SOURCE_LIMIT = 1
ANSWER_FORMATS = ("STAR", "XYZ")

_TITLE_NOISE = frozenset(
    {
        "assistant",
        "associate",
        "chief",
        "director",
        "executive",
        "head",
        "junior",
        "lead",
        "manager",
        "principal",
        "senior",
        "snr",
        "staff",
        "vice",
        "vp",
    }
)

_OCCUPATION_GENERIC_TERMS = frozenset({"engineer"})


class ResearchAccessFailure(RuntimeError):
    """An approved source could not be reached or parsed."""


@dataclass(frozen=True)
class ResearchPack:
    status: Literal["complete", "partial", "valid_empty", "access_failure"]
    role_company_brief: dict
    interview_pack: dict
    compensation_brief: dict
    source_statuses: tuple[dict, ...]
    built_at: str


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _normalized_words(value: str) -> list[str]:
    words = re.findall(r"[a-z0-9+#.]+", (value or "").casefold())
    meaningful = [word for word in words if len(word) > 1 and word not in _TITLE_NOISE]
    return meaningful or words


def _source_age_state(value: str, now: datetime) -> str:
    try:
        parsed = datetime.fromisoformat((value or "").replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
    except ValueError:
        return "unknown"
    return "stale" if (now - parsed).days > SOURCE_STALE_AFTER_DAYS else "fresh"


def _job_source(job: ScrapedJob, *, confidence: str, note: str, now: datetime) -> dict:
    return {
        "url": job.url or "",
        "publisher": job.source or "",
        "source_type": "job_posting",
        "retrieved_at": job.scraped_at or "",
        "data_date": job.posted_date or "",
        "freshness": _source_age_state(job.scraped_at or job.posted_at_sort or "", now),
        "confidence": confidence,
        "evidence_note": note,
    }


def _resume_evidence(resume_text: str, term: str) -> str:
    terms = set(_normalized_words(term))
    if not terms:
        return ""
    best_line = ""
    best_overlap = 0
    for raw in (resume_text or "").splitlines():
        line = raw.strip(" \t-•")
        if not MIN_EVIDENCE_LINE_CHARS <= len(line) <= MAX_EVIDENCE_LINE_CHARS:
            continue
        overlap = len(terms.intersection(_normalized_words(line)))
        if overlap > best_overlap:
            best_line = line
            best_overlap = overlap
    return best_line if best_overlap else ""


def _answer_scaffold(question: str, evidence: str) -> dict:
    if not evidence:
        return {
            "status": "missing_evidence",
            "evidence_quote": "",
            "steps": [
                "Confirm a real example before answering.",
                "State the situation and your responsibility.",
                "Describe only actions and results you can defend.",
            ],
        }
    return {
        "status": "evidence_ready",
        "evidence_quote": evidence,
        "steps": [
            "Situation and task: add the context around the cited resume evidence.",
            "Action: explain your own decisions and trade-offs.",
            "Result and reflection: use only verified outcomes, then state what you learned.",
        ],
        "question": question,
    }


def _occupation_score(role: str, occupation: str) -> float:
    role_words = set(_normalized_words(role)) - _OCCUPATION_GENERIC_TERMS
    occupation_words = set(_normalized_words(occupation)) - _OCCUPATION_GENERIC_TERMS
    if not role_words or not occupation_words:
        return 0.0
    intersection = len(role_words & occupation_words)
    union = len(role_words | occupation_words)
    jaccard = intersection / union if union else 0.0
    sequence = SequenceMatcher(
        None,
        " ".join(sorted(role_words)),
        " ".join(sorted(occupation_words)),
    ).ratio()
    subset_bonus = OCCUPATION_SUBSET_BONUS if occupation_words <= role_words or role_words <= occupation_words else 0.0
    return min(
        1.0,
        (OCCUPATION_JACCARD_WEIGHT * jaccard) + (OCCUPATION_SEQUENCE_WEIGHT * sequence) + subset_bonus,
    )


def _mom_observation(role: str, response_content: bytes) -> dict | None:
    try:
        workbook = load_workbook(BytesIO(response_content), read_only=True, data_only=True)
        sheet = workbook[MOM_WAGE_SHEET]
    except Exception as error:
        raise ResearchAccessFailure("MOM wage workbook could not be parsed") from error

    candidates: list[tuple[float, tuple]] = []
    for row in sheet.iter_rows(min_row=MOM_WAGE_FIRST_DATA_ROW, values_only=True):
        occupation = row[2] if len(row) > 2 else None
        if not isinstance(occupation, str) or not occupation.strip():
            continue
        wages = row[3:9]
        if len(wages) != 6 or not all(isinstance(value, (int, float)) for value in wages):
            continue
        candidates.append((_occupation_score(role, occupation), row))
    if not candidates:
        return None
    score, best = max(candidates, key=lambda item: item[0])
    if score < MOM_WAGE_MATCH_MIN_SCORE:
        return None
    return {
        "source_url": MOM_WAGE_LANDING_URL,
        "source_type": "government_statistics",
        "publisher": "Singapore Ministry of Manpower",
        "retrieved_at": _utcnow().isoformat(),
        "data_date": MOM_WAGE_DATA_PERIOD,
        "release_date": MOM_WAGE_RELEASE_DATE,
        "occupation": best[2],
        "industry": "All industries",
        "population": "Full-time resident employees",
        "currency": "SGD",
        "period": "monthly",
        "basic_wage": {"p25": best[3], "median": best[4], "p75": best[5]},
        "gross_wage": {"p25": best[6], "median": best[7], "p75": best[8]},
        "excludes": ["bonuses", "stock options", "employer CPF"],
        "role_mapping": {
            "target_title": role,
            "matched_occupation": best[2],
            "match_score": round(score, OCCUPATION_MATCH_SCORE_DECIMALS),
            "context": "Closest SSOC 2024 occupation label in the all-industries table.",
        },
    }


class CorpusAndMomResearchProvider:
    """Approved public-job corpus plus the current official MOM wage release."""

    def __init__(
        self,
        db: Session,
        http_get: Callable[..., requests.Response] = requests.get,
    ) -> None:
        self._db = db
        self._http_get = http_get

    def build(self, tracked: TrackedJob, resume_text: str) -> ResearchPack:
        now = _utcnow()
        target = self._target_job(tracked)
        comparable = self._comparable_jobs(tracked, target)
        company_jobs = self._company_jobs(tracked, target)
        job_sources = self._job_sources(tracked, target, comparable, company_jobs, now)
        role_company_brief, ats_terms = self._role_company_brief(
            tracked,
            target,
            comparable,
            company_jobs,
            job_sources,
            now,
        )
        interview_pack = self._interview_pack(
            tracked,
            ats_terms,
            resume_text,
            job_sources,
            role_company_brief["freshness"],
        )
        compensation_brief, wage_status = self._compensation_brief(tracked, target, now)
        source_statuses = (
            {
                "source": "public_job_corpus",
                "status": "complete" if job_sources else "valid_empty",
                "result_count": len(job_sources),
            },
            wage_status,
            {
                "source": "community_and_employer_reviews",
                "status": "valid_empty",
                "result_count": 0,
                "source_type": "restricted_review_sources",
                "retrieved_at": now.isoformat(),
                "confidence": "none",
                "detail": "No approved public adapter is configured; restricted sources were not scraped.",
            },
        )
        successful = sum(item["status"] == "complete" for item in source_statuses)
        failed = sum(item["status"] == "access_failure" for item in source_statuses)
        if successful and failed:
            status: Literal["complete", "partial", "valid_empty", "access_failure"] = "partial"
        elif successful:
            status = "complete"
        elif failed:
            status = "access_failure"
        else:
            status = "valid_empty"
        return ResearchPack(
            status=status,
            role_company_brief=role_company_brief,
            interview_pack=interview_pack,
            compensation_brief=compensation_brief,
            source_statuses=source_statuses,
            built_at=now.isoformat(),
        )

    def _target_job(self, tracked: TrackedJob) -> ScrapedJob | None:
        if tracked.scraped_job_id is None:
            return None
        return self._db.get(ScrapedJob, tracked.scraped_job_id)

    def _comparable_jobs(
        self,
        tracked: TrackedJob,
        target: ScrapedJob | None,
    ) -> list[ScrapedJob]:
        words = _normalized_words(tracked.role)[:TITLE_QUERY_TERM_LIMIT]
        if not words:
            return []
        query = apply_public_job_visibility(self._db.query(ScrapedJob)).filter(
            or_(*(ScrapedJob.title.ilike(f"%{word}%") for word in words))
        )
        if target is not None:
            query = query.filter(ScrapedJob.id != target.id)
        return query.order_by(ScrapedJob.posted_at_sort.desc(), ScrapedJob.id.desc()).limit(COMPARABLE_JOB_LIMIT).all()

    def _company_jobs(
        self,
        tracked: TrackedJob,
        target: ScrapedJob | None,
    ) -> list[ScrapedJob]:
        query = apply_public_job_visibility(self._db.query(ScrapedJob)).filter(ScrapedJob.company == tracked.company)
        if target is not None:
            query = query.filter(ScrapedJob.id != target.id)
        return query.order_by(ScrapedJob.posted_at_sort.desc(), ScrapedJob.id.desc()).limit(COMPANY_JOB_LIMIT).all()

    @staticmethod
    def _job_sources(
        tracked: TrackedJob,
        target: ScrapedJob | None,
        comparable: list[ScrapedJob],
        company_jobs: list[ScrapedJob],
        now: datetime,
    ) -> list[dict]:
        sources: list[dict] = []
        seen: set[str] = set()
        if target is not None:
            sources.append(
                _job_source(
                    target,
                    confidence="high",
                    note="Exact selected posting snapshot.",
                    now=now,
                )
            )
            seen.add(target.url or f"job:{target.id}")
        elif tracked.source_url:
            sources.append(
                {
                    "url": tracked.source_url,
                    "publisher": tracked.source or "",
                    "source_type": "job_posting",
                    "retrieved_at": "",
                    "data_date": "",
                    "freshness": "unknown",
                    "confidence": "high",
                    "evidence_note": "Saved posting snapshot; live source metadata was unavailable.",
                }
            )
            seen.add(tracked.source_url)
        for job in [*comparable, *company_jobs]:
            key = job.url or f"job:{job.id}"
            if key in seen:
                continue
            seen.add(key)
            sources.append(
                _job_source(
                    job,
                    confidence="medium",
                    note=(
                        "Current posting from the same company."
                        if job.company == tracked.company
                        else "Current comparable role posting."
                    ),
                    now=now,
                )
            )
        return sources

    def _role_company_brief(
        self,
        tracked: TrackedJob,
        target: ScrapedJob | None,
        comparable: list[ScrapedJob],
        company_jobs: list[ScrapedJob],
        sources: list[dict],
        now: datetime,
    ) -> tuple[dict, list[dict]]:
        source_by_url = {item["url"]: item for item in sources}
        term_rows: dict[str, dict] = {}
        occurrences: Counter[str] = Counter()
        source_urls: defaultdict[str, list[str]] = defaultdict(list)
        jobs = ([target] if target is not None else []) + comparable
        if target is None and tracked.job_description:
            synthetic_terms = build_job_ats_terms(
                tracked.job_description,
                job_title=tracked.role,
                limit=ATS_TERM_LIMIT,
                db_session=self._db,
            )
            jobs_with_terms = [(None, synthetic_terms)]
        else:
            jobs_with_terms = [
                (
                    job,
                    build_job_ats_terms(
                        job.description or "",
                        job.skills if isinstance(job.skills, list) else [],
                        job.parsed_jd if isinstance(job.parsed_jd, dict) else None,
                        job.title or "",
                        limit=ATS_TERM_LIMIT,
                        db_session=self._db,
                    ),
                )
                for job in jobs
                if job is not None
            ]
        for job, terms in jobs_with_terms:
            for term in terms:
                key = str(term.get("skill") or "").strip().casefold()
                if not key:
                    continue
                occurrences[key] += 1
                if job is not None and job.url and job.url not in source_urls[key]:
                    source_urls[key].append(job.url)
                elif job is None and sources and sources[0]["url"] not in source_urls[key]:
                    source_urls[key].append(sources[0]["url"])
                if key not in term_rows or term.get("required"):
                    term_rows[key] = term
        ordered_terms = sorted(
            term_rows,
            key=lambda key: (
                -int(bool(term_rows[key].get("required"))),
                -occurrences[key],
                key,
            ),
        )[:ATS_TERM_LIMIT]
        ats_terms = [
            {
                "term": key,
                "context": term_rows[key].get("jd_context") or "",
                "confidence": (
                    "high"
                    if term_rows[key].get("required") or occurrences[key] >= RECURRING_TERM_MIN_POSTINGS
                    else "medium"
                ),
                "observed_in_postings": occurrences[key],
                "sources": [source_by_url[url] for url in source_urls[key] if url in source_by_url],
            }
            for key in ordered_terms
        ]
        comparable_titles = [
            {
                "title": job.title,
                "company": job.company,
                "source": source_by_url.get(job.url),
            }
            for job in comparable
        ]
        industries = sorted(
            {
                value
                for job in ([target] if target is not None else []) + company_jobs
                for value in (job.sector, job.company_ssic_description)
                if value
            }
        )
        return {
            "status": "complete" if sources else "valid_empty",
            "company": {
                "name": tracked.company,
                "current_posting_count": len(company_jobs) + (1 if target is not None else 0),
                "observed_industries": industries,
                "other_current_titles": [job.title for job in company_jobs],
                "context": "Observed only from attributable current public job postings.",
            },
            "role": {
                "title": tracked.role,
                "comparable_titles": comparable_titles,
                "ats_terms": ats_terms,
            },
            "sources": sources,
            "freshness": (
                "stale"
                if sources and all(item.get("freshness") == "stale" for item in sources)
                else "fresh"
                if any(item.get("freshness") == "fresh" for item in sources)
                else "unknown"
            ),
            "researched_at": now.isoformat(),
        }, ats_terms

    @staticmethod
    def _interview_pack(
        tracked: TrackedJob,
        ats_terms: list[dict],
        resume_text: str,
        sources: list[dict],
        source_freshness: str,
    ) -> dict:
        questions: list[dict] = []
        for term in ats_terms[:INTERVIEW_TERM_LIMIT]:
            label = term["term"]
            evidence = _resume_evidence(resume_text, label)
            question = f"Tell me about a time you applied {label} in work comparable to this role."
            questions.append(
                {
                    "cluster": "technical",
                    "question": question,
                    "confidence": term["confidence"],
                    "source_type": "job_posting",
                    "sources": term.get("sources", []),
                    "answer_scaffold": _answer_scaffold(question, evidence),
                }
            )
        general = [
            (
                "behavioral",
                "Describe a cross-functional disagreement you had to resolve and what changed afterward.",
                "medium",
            ),
            (
                "company",
                f"Why {tracked.company}, and what in its current role context matters to you?",
                "high",
            ),
            (
                "role",
                f"What would you prioritise in your first 90 days as {tracked.role}?",
                "high",
            ),
            (
                "recruiter_screen",
                "Which parts of the role, package, working model, and timeline do you need clarified?",
                "high",
            ),
        ]
        for cluster, question, confidence in general:
            evidence = _resume_evidence(resume_text, tracked.role) if cluster == "role" else ""
            question_sources = sources[:PRIMARY_SOURCE_LIMIT]
            questions.append(
                {
                    "cluster": cluster,
                    "question": question,
                    "confidence": confidence,
                    "source_type": "job_posting" if question_sources else "candidate_context",
                    "sources": question_sources,
                    "answer_scaffold": (
                        _answer_scaffold(question, evidence)
                        if cluster in {"behavioral", "role"}
                        else {"status": "candidate_input_required", "evidence_quote": "", "steps": []}
                    ),
                }
            )
        return {
            "status": "complete" if sources else "sparse",
            "source_state": source_freshness if sources else "valid_empty",
            "questions": questions,
            "source_leads": sources,
            "answer_formats": list(ANSWER_FORMATS),
            "confidence_note": (
                "These are evidence-grounded preparation clusters, not guaranteed interview questions."
            ),
            "restricted_source_status": "not_searched_without_an_approved_adapter",
        }

    def _compensation_brief(
        self,
        tracked: TrackedJob,
        target: ScrapedJob | None,
        now: datetime,
    ) -> tuple[dict, dict]:
        observations: list[dict] = []
        pipeline = (tracked.role_metadata or {}).get("recruitment_pipeline") or {}
        posting = pipeline.get("posting_snapshot") if isinstance(pipeline, dict) else {}
        posting_salary = display_salary(
            str((target.salary if target is not None else "") or (posting or {}).get("salary") or "")
        )
        if posting_salary:
            source = (posting or {}).get("source") if isinstance(posting, dict) else {}
            observations.append(
                {
                    "kind": "employer_posting",
                    "value": posting_salary,
                    "currency": "SGD",
                    "period": "as stated by employer",
                    "definition": "Employer-stated posting range; package components were not inferred.",
                    "source_url": (
                        (target.url if target is not None else "")
                        or (source or {}).get("url")
                        or tracked.source_url
                        or ""
                    ),
                    "source_type": "job_posting",
                    "data_date": (
                        (target.posted_date if target is not None else "") or (source or {}).get("posted_date") or ""
                    ),
                }
            )
        try:
            response = self._http_get(MOM_WAGE_SOURCE_URL, timeout=MOM_WAGE_TIMEOUT_SECONDS)
            response.raise_for_status()
            mom = _mom_observation(tracked.role, response.content)
            if mom is not None:
                observations.append({"kind": "mom_occupational_wages", **mom})
                wage_status = {
                    "source": "mom_occupational_wages_2025",
                    "status": "complete",
                    "result_count": 1,
                }
            else:
                wage_status = {
                    "source": "mom_occupational_wages_2025",
                    "status": "valid_empty",
                    "result_count": 0,
                    "detail": "No sufficiently similar SSOC occupation was found.",
                }
        except (requests.RequestException, ResearchAccessFailure) as error:
            wage_status = {
                "source": "mom_occupational_wages_2025",
                "status": "access_failure",
                "result_count": 0,
                "detail": str(error) or type(error).__name__,
            }
        return {
            "status": "complete" if observations else wage_status["status"],
            "observations": observations,
            "recruiter_guide_leads": [
                {
                    "publisher": "Hays Singapore",
                    "title": "2026 Asia Salary Guide",
                    "source_url": HAYS_SALARY_GUIDE_URL,
                    "source_type": "recruiter_guide",
                    "publication_date": HAYS_SALARY_GUIDE_PUBLICATION_DATE,
                    "retrieved_at": now.isoformat(),
                    "status": "source_lead",
                    "evidence_note": (
                        "Public guide landing page; no numeric range is copied unless its role and "
                        "package definition are accessible and compatible."
                    ),
                },
                {
                    "publisher": "Michael Page Singapore",
                    "title": "Salary benchmark tool",
                    "source_url": MICHAEL_PAGE_SALARY_BENCHMARK_URL,
                    "source_type": "recruiter_guide",
                    "publication_date": "not stated",
                    "retrieved_at": now.isoformat(),
                    "status": "source_lead",
                    "evidence_note": ("Public benchmark landing page; interactive or gated figures are not scraped."),
                },
            ],
            "comparison_state": (
                "multiple_incompatible_observations"
                if len(observations) > 1
                else "sparse"
                if len(observations) == 1
                else "valid_empty"
            ),
            "comparison_rule": (
                "Monthly basic, monthly gross, posting range, bonus, benefits, and total package "
                "remain separate observations and are never silently averaged."
            ),
            "restricted_source_status": "Glassdoor and other restricted sources were not scraped.",
        }, wage_status
